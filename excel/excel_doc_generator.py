from datetime import datetime
import re
from typing import List

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from extraction.pub_extract import PubExtract
from meeting.meeting_section import MeetingSection
from meeting.section_kind import SectionKind

from .config import ExcelConfig


class ExcelGenerator(object):

    def __init__(self, publication_extracts: List[PubExtract], labels: dict, config: ExcelConfig):
        self.publication_extracts = publication_extracts
        self.labels = labels
        self.__CURRENT_ROW = 3
        self.__LEFT_COLUMNS = ['B', 'C', 'D', 'E']
        self.__RIGHT_COLUMNS = ['G', 'H', 'I', 'J']
        self.__ACTIVE_COLUMNS = self.__LEFT_COLUMNS
        self.config = config
        # a new sheet is created for each publication so the program won't
        # use the first sheet (it will be empty). Hence, it it removed
        self.workbook = Workbook()
        self.workbook.remove_sheet(worksheet=self.workbook.get_active_sheet())

    def create_excel_doc(self):
        if len(self.publication_extracts) == 0:
            return
        print('creating excel document...')

        for publication_extract in self.publication_extracts:
            print("preparing mwb({}) {}/{}".format(
                self.labels['lang_key'], publication_extract.pub_month, publication_extract.pub_year), end='  ')
            self._add_populated_sheet(publication_extract)
            print('[{}OK{}]'.format('\033[92m', '\033[0m'))

        file_name = 'WREX_{}_{}.xlsx'.format(datetime.now().strftime("%m-%d-%Y_%H:%M"), self.labels['lang_key'])
        self.workbook.save(file_name)
        print('done...', "saved to '{}'".format(file_name))

    def _add_populated_sheet(self, publication_extract: PubExtract):
        sheet = self.workbook.create_sheet(publication_extract.pub_name)
        meetings_count = 0

        self._insert_sheet_title(
            sheet,
            self.labels[publication_extract.pub_month] + publication_extract.pub_year)

        for meeting in publication_extract.meetings:
            self._insert_header_content(meeting.week_span, sheet)
            self._insert_section(meeting.treasures_section, sheet)
            self._insert_section(meeting.ministry_section, sheet)
            self._insert_section(meeting.christian_section, sheet)
            self._insert_footer_content(sheet)

            meetings_count += 1
            if meetings_count == 3:
                self.__CURRENT_ROW = 5
                self.__ACTIVE_COLUMNS = self.__RIGHT_COLUMNS
            self._final_styling_touches(sheet)
        # reset indexes and make them ready for the next sheet
        self.__CURRENT_ROW = 3
        self.__ACTIVE_COLUMNS = self.__LEFT_COLUMNS

    def _insert_sheet_title(self, sheet: Worksheet, month_name_and_year: str):
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__LEFT_COLUMNS[0] + str(current_row)
        end_cell = self.__RIGHT_COLUMNS[3] + str(current_row)
        sheet[start_cell] = self.labels['meeting_name'] + ' – ' + month_name_and_year
        sheet.merge_cells(start_cell + ':' + end_cell)
        self.__CURRENT_ROW += 2
        self._style_sheet_title(sheet)

    def _style_sheet_title(self, sheet: Worksheet):
        cell = self.__ACTIVE_COLUMNS[0] + str(3)
        sheet[cell].font = Font(bold=True, size=self.config.SHEET_TITLE_FONT_SIZE)
        sheet[cell].alignment = Alignment(horizontal='center', vertical='center')

    def _insert_header_content(self, week_span: str, sheet: Worksheet):
        # week span
        week_span_row = self.__CURRENT_ROW
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[0] + current_row
        end_cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[start_cell] = re.sub('[-|–]', ' – ', week_span)
        sheet.merge_cells(start_cell + ':' + end_cell)
        self.__CURRENT_ROW += 1
        # chairman
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[0] + current_row
        end_cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[start_cell] = self.labels['chairman']
        sheet.merge_cells(start_cell + ':' + end_cell)
        self.__CURRENT_ROW += 1
        # opening prayer
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[start_cell] = self.labels['opening_prayer']
        self.__CURRENT_ROW += 1
        self._style_header_content(week_span_row, sheet)

    def _style_header_content(self, week_span_row: int, sheet: Worksheet):
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(bottom=Side(border_style='thin'))
        # week span
        cell = self.__ACTIVE_COLUMNS[0] + str(week_span_row)
        sheet[cell].font = Font(bold=True, size=self.config.LARGE_FONT_SIZE)
        sheet[cell].alignment = alignment
        # chairman
        cell = self.__ACTIVE_COLUMNS[0] + str(week_span_row + 1)
        sheet[cell].font = Font(bold=True, size=self.config.LARGE_FONT_SIZE)
        sheet[cell].alignment = alignment
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(week_span_row + 1)  # last cell
        sheet[cell].border = border
        # opening prayer
        cell = self.__ACTIVE_COLUMNS[2] + str(week_span_row + 2)
        sheet[cell].font = Font(size=self.config.SMALL_FONT_SIZE)
        sheet[cell].alignment = alignment
        cell = self.__ACTIVE_COLUMNS[3] + str(week_span_row + 2)  # last cell
        sheet[cell].border = border

    def _insert_section_title(self, meeting_section: MeetingSection, sheet: Worksheet):
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[0] + current_row
        sheet[start_cell] = meeting_section.title
        # by default the right most cell at which merging cells ends is on the 5th column (including the offset)
        end_cell = self.__ACTIVE_COLUMNS[3] + current_row
        # but if the section is 'IMPROVE_IN_MINISTRY' and the user requests hall-dividing
        # rows then it changes to the 3rd row
        if (meeting_section.section_kind == SectionKind.IMPROVE_IN_MINISTRY) and \
                self.config.INSERT_HALL_DIVISION_LABELS:
            end_cell = self.__ACTIVE_COLUMNS[1] + current_row
            self._insert_hall_divider(sheet, meeting_section.section_kind)

        sheet.merge_cells(start_cell + ':' + end_cell)
        self._style_section_title(self.__CURRENT_ROW, sheet, meeting_section.section_kind)
        self.__CURRENT_ROW += 1

    def _style_section_title(self, title_row: int, sheet: Worksheet, section_kind: SectionKind):
        if section_kind == SectionKind.TREASURES:
            bg_color = self.config.TREASURES_SECTION_TITLE_BG_COLOR
        elif section_kind == SectionKind.IMPROVE_IN_MINISTRY:
            bg_color = self.config.MINISTRY_SECTION_TITLE_BG_COLOR
        else:
            bg_color = self.config.CHRISTIAN_LIFE_SECTION_TITLE_BG_COLOR

        cell = self.__ACTIVE_COLUMNS[0] + str(title_row)
        sheet[cell].font = Font(bold=True, size=self.config.LARGE_FONT_SIZE)
        sheet[cell].alignment = Alignment(horizontal='left', vertical='center')
        sheet[cell].fill = PatternFill(patternType='solid', fgColor=bg_color)
        sheet[cell].border = Border(top=Side(border_style='thin'))

    def _insert_hall_divider(self, sheet, section_kind: SectionKind):
        current_row = str(self.__CURRENT_ROW)
        main_hall_cell = self.__ACTIVE_COLUMNS[2] + current_row
        second_hall_cell = self.__ACTIVE_COLUMNS[3] + current_row
        sheet[main_hall_cell] = self.labels['main_hall']
        sheet[second_hall_cell] = self.labels['second_hall']
        self._style_hall_divider(self.__CURRENT_ROW, sheet, section_kind)

    def _style_hall_divider(self, divider_row: int, sheet: Worksheet, section_kind: SectionKind):
        main_hall_cell = self.__ACTIVE_COLUMNS[2] + str(divider_row)
        second_hall_cell = self.__ACTIVE_COLUMNS[3] + str(divider_row)
        font = Font(bold=False, size=self.config.SMALL_FONT_SIZE)
        alignment = Alignment(horizontal='center', vertical='center')
        if section_kind == SectionKind.IMPROVE_IN_MINISTRY:
            fg_color = self.config.MINISTRY_SECTION_TITLE_BG_COLOR
        else:
            fg_color = self.config.DEFAULT_BG_COLOR
        fill = PatternFill(patternType='solid', fgColor=fg_color)
        border = Border(top=Side(border_style='thin'))

        sheet[main_hall_cell].font = font
        sheet[main_hall_cell].alignment = alignment
        sheet[main_hall_cell].fill = fill
        sheet[main_hall_cell].border = border
        sheet[second_hall_cell].font = font
        sheet[second_hall_cell].alignment = alignment
        sheet[second_hall_cell].fill = fill
        sheet[second_hall_cell].border = border

    def _insert_section(self, meeting_section: MeetingSection, sheet: Worksheet):
        self._insert_section_title(meeting_section, sheet)
        bible_reading = self.labels['bible_reading']

        for presentation in meeting_section.presentations:
            if meeting_section.section_kind == SectionKind.TREASURES:
                if (bible_reading in presentation) and self.config.INSERT_HALL_DIVISION_LABELS:
                    self._insert_hall_divider(sheet, meeting_section.section_kind)
                    self.__CURRENT_ROW += 1
            current_row = str(self.__CURRENT_ROW)
            start_cell = self.__ACTIVE_COLUMNS[1] + current_row

            if (meeting_section.section_kind == SectionKind.IMPROVE_IN_MINISTRY) and\
                    self.config.INSERT_HALL_DIVISION_LABELS:
                end_cell = self.__ACTIVE_COLUMNS[1] + current_row
            else:
                end_cell = self.__ACTIVE_COLUMNS[2] + current_row
            sheet[start_cell] = presentation

            if (bible_reading not in presentation) or \
                    (bible_reading in presentation and not self.config.INSERT_HALL_DIVISION_LABELS):
                sheet.merge_cells(start_cell + ':' + end_cell)
            self._style_section(self.__CURRENT_ROW, sheet)
            self.__CURRENT_ROW += 1

    def _style_section(self, presentation_row: int, sheet: Worksheet):
        cell = self.__ACTIVE_COLUMNS[1] + str(presentation_row)
        border = Border(bottom=Side(border_style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')
        sheet[cell].font = Font(size=self.config.LARGE_FONT_SIZE)
        sheet[cell].alignment = Alignment(horizontal='left', vertical='center')
        sheet[cell].border = border
        # underline and center the last two cells of a presentation row (that is where presenter names go)
        cell = self.__ACTIVE_COLUMNS[2] + str(presentation_row)
        sheet[cell].border = border
        sheet[cell].alignment = alignment
        cell = self.__ACTIVE_COLUMNS[3] + str(presentation_row)
        sheet[cell].border = border
        sheet[cell].alignment = alignment

    def _insert_footer_content(self, sheet: Worksheet):
        footer_row = self.__CURRENT_ROW
        current_row = str(self.__CURRENT_ROW)
        cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[cell] = self.labels['reader']
        self.__CURRENT_ROW += 1

        current_row = str(self.__CURRENT_ROW)
        cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[cell] = self.labels['concluding_prayer']
        self.__CURRENT_ROW += 3
        self._style_footer_content(footer_row, sheet)

    def _style_footer_content(self, footer_row: int, sheet: Worksheet):
        font = Font(size=self.config.SMALL_FONT_SIZE)
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(bottom=Side(border_style='thin'))

        cell = self.__ACTIVE_COLUMNS[2] + str(footer_row)
        sheet[cell].font = font
        sheet[cell].alignment = alignment
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(footer_row)
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[2] + str(footer_row + 1)
        sheet[cell].font = font
        sheet[cell].alignment = alignment
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(footer_row + 1)
        sheet[cell].border = border

    def _final_styling_touches(self, sheet: Worksheet):
        sheet.page_setup.paperSize = Worksheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins = PageMargins(left=self.config.MARGIN_LENGTH, right=self.config.MARGIN_LENGTH)
        sheet.column_dimensions[self.__LEFT_COLUMNS[0]].width = 4
        sheet.column_dimensions[self.__LEFT_COLUMNS[1]].width = 42
        sheet.column_dimensions[self.__RIGHT_COLUMNS[0]].width = 4
        sheet.column_dimensions[self.__RIGHT_COLUMNS[1]].width = 42

        for rows in sheet.iter_rows(min_row=3):
            for row in rows:
                sheet.row_dimensions[row.row].height = self.config.ROW_HEIGHT
