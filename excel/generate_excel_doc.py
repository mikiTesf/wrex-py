from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from datetime import datetime

from typing import List
from meeting.meeting import Meeting
from meeting.meeting_section import MeetingSection
from meeting.section_kind import SectionKind


class ExcelGenerator(object):

    def __init__(self, meeting_objects: List[List[Meeting]]):
        self.workbook = Workbook()
        self.meeting_object_lists = meeting_objects
        self.__CURRENT_ROW = 3
        self.__LEFT_COLUMNS = ['B', 'C', 'D', 'E']
        self.__RIGHT_COLUMNS = ['G', 'H', 'I', 'J']
        self.__ACTIVE_COLUMNS = self.__LEFT_COLUMNS
        # a new sheet is created for each publication so the program won't
        # use the first sheet (it will be empty). Hence, it it removed
        self.workbook.remove_sheet(worksheet=self.workbook.get_active_sheet())
        # formatting constants
        self.SMALL_FONT_SIZE = 15
        self.LARGE_FONT_SIZE = 16

    def create_excel_doc(self):
        print('creating excel document...')

        for meeting_object_list in self.meeting_object_lists:
            self._add_populated_sheet(meeting_object_list)

        file_name = 'wrex {}.xlsx'.format(datetime.now().strftime("%m-%d-%Y_%H:%M"))
        self.workbook.save(file_name)
        print('done...')

    def _add_populated_sheet(self, meeting_objects: List[Meeting]):
        sheet = self.workbook.create_sheet('pub')  # todo: change with publication `file name`
        meetings_count = 0

        self._insert_sheet_title(sheet)
        for meeting in meeting_objects:
            self._insert_header_content(meeting.week_span, sheet)
            self._insert_section(meeting.treasures_section, sheet)
            self._insert_section(meeting.ministry_section, sheet)
            self._insert_section(meeting.christian_section, sheet)
            self._insert_footer_content(sheet)

            meetings_count += 1
            if meetings_count == 3:
                self.__CURRENT_ROW = 5
                self.__ACTIVE_COLUMNS = self.__RIGHT_COLUMNS
            self._final_styling_touches(sheet)
        # reset indexes and make them ready for the next sheet
        self.__CURRENT_ROW = 3
        self.__ACTIVE_COLUMNS = self.__LEFT_COLUMNS

    def _insert_sheet_title(self, sheet: Worksheet):
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__LEFT_COLUMNS[0] + str(current_row)
        end_cell = self.__RIGHT_COLUMNS[3] + str(current_row)
        sheet[start_cell] = 'Christian Life And Ministry'  # todo: get value from config file
        sheet.merge_cells(start_cell + ':' + end_cell)
        self.__CURRENT_ROW += 2
        self._style_sheet_title(sheet)

    def _style_sheet_title(self, sheet: Worksheet):
        cell = self.__ACTIVE_COLUMNS[0] + str(3)
        sheet[cell].font = Font(bold=True, size=self.LARGE_FONT_SIZE)
        sheet[cell].alignment = Alignment(horizontal='center', vertical='center')

    def _insert_header_content(self, week_span: str, sheet: Worksheet):
        # week span
        week_span_row = self.__CURRENT_ROW
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[0] + current_row
        end_cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[start_cell] = week_span  # todo: get month name from config file
        sheet.merge_cells(start_cell + ':' + end_cell)
        self.__CURRENT_ROW += 1
        # chairman
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[0] + current_row
        end_cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[start_cell] = 'Chairman'  # todo: get equivalent from config file
        sheet.merge_cells(start_cell + ':' + end_cell)
        self.__CURRENT_ROW += 1
        # opening prayer
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[start_cell] = 'Opening Prayer'  # todo: get equivalent from config file
        self.__CURRENT_ROW += 1
        self._style_header_content(week_span_row, sheet)

    def _style_header_content(self, week_span_row: int, sheet: Worksheet):
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(bottom=Side(border_style='thin'))
        # week span
        cell = self.__ACTIVE_COLUMNS[0] + str(week_span_row)
        sheet[cell].font = Font(bold=True, size=self.LARGE_FONT_SIZE)
        sheet[cell].alignment = alignment
        # chairman
        cell = self.__ACTIVE_COLUMNS[0] + str(week_span_row + 1)
        sheet[cell].font = Font(bold=True, size=self.LARGE_FONT_SIZE)
        sheet[cell].alignment = alignment
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(week_span_row + 1)  # last cell
        sheet[cell].border = border
        # opening prayer
        cell = self.__ACTIVE_COLUMNS[2] + str(week_span_row + 2)
        sheet[cell].font = Font(size=self.SMALL_FONT_SIZE)
        sheet[cell].alignment = alignment
        cell = self.__ACTIVE_COLUMNS[3] + str(week_span_row + 2)  # last cell
        sheet[cell].border = border

    def _insert_section_title(self, meeting_section: MeetingSection, sheet: Worksheet):
        current_row = str(self.__CURRENT_ROW)
        start_cell = self.__ACTIVE_COLUMNS[0] + current_row
        sheet[start_cell] = meeting_section.title

        if meeting_section.section_kind == SectionKind.IMPROVE_IN_MINISTRY:
            end_cell = self.__ACTIVE_COLUMNS[1] + current_row
            self._insert_hall_divider(sheet)
        else:
            end_cell = self.__ACTIVE_COLUMNS[3] + current_row
        sheet.merge_cells(start_cell + ':' + end_cell)

        self._style_section_title(self.__CURRENT_ROW, sheet)
        self.__CURRENT_ROW += 1

    def _style_section_title(self, title_row: int, sheet: Worksheet):
        cell = self.__ACTIVE_COLUMNS[0] + str(title_row)
        sheet[cell].font = Font(bold=True, size=self.LARGE_FONT_SIZE)
        sheet[cell].alignment = Alignment(horizontal='left', vertical='center')
        sheet[cell].fill = PatternFill(patternType='solid', fgColor='C8C8C8')
        sheet[cell].border = Border(top=Side(border_style='thin'))

    def _insert_hall_divider(self, sheet):
        current_row = str(self.__CURRENT_ROW)
        main_hall_cell = self.__ACTIVE_COLUMNS[2] + current_row
        second_hall_cell = self.__ACTIVE_COLUMNS[3] + current_row
        sheet[main_hall_cell] = 'Main Hall'  # todo: get value from config file
        sheet[second_hall_cell] = 'Second Hall'  # todo: get value from config file
        self._style_hall_divider(self.__CURRENT_ROW, sheet)

    def _style_hall_divider(self, divider_row: int, sheet: Worksheet):
        main_hall_cell = self.__ACTIVE_COLUMNS[2] + str(divider_row)
        second_hall_cell = self.__ACTIVE_COLUMNS[3] + str(divider_row)
        font = Font(bold=False, size=self.SMALL_FONT_SIZE)
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(patternType='solid', fgColor='C8C8C8')
        border = Border(top=Side(border_style='thin'))

        sheet[main_hall_cell].font = font
        sheet[main_hall_cell].alignment = alignment
        sheet[main_hall_cell].fill = fill
        sheet[main_hall_cell].border = border
        sheet[second_hall_cell].font = font
        sheet[second_hall_cell].alignment = alignment
        sheet[second_hall_cell].fill = fill
        sheet[second_hall_cell].border = border

    def _insert_section(self, meeting_section: MeetingSection, sheet: Worksheet):
        self._insert_section_title(meeting_section, sheet)
        bible_reading = 'Bible Reading:'  # todo: get value from config file

        for presentation in meeting_section.presentations:
            if meeting_section.section_kind == SectionKind.TREASURES:
                if bible_reading in presentation:
                    self._insert_hall_divider(sheet)
                    self.__CURRENT_ROW += 1
            current_row = str(self.__CURRENT_ROW)
            start_cell = self.__ACTIVE_COLUMNS[1] + current_row

            if meeting_section.section_kind == SectionKind.IMPROVE_IN_MINISTRY:
                end_cell = self.__ACTIVE_COLUMNS[1] + current_row
            else:
                end_cell = self.__ACTIVE_COLUMNS[2] + current_row
            sheet[start_cell] = presentation

            if bible_reading not in presentation:
                sheet.merge_cells(start_cell + ':' + end_cell)
            self._style_section(self.__CURRENT_ROW, sheet)
            self.__CURRENT_ROW += 1

    def _style_section(self, presentation_row: int, sheet: Worksheet):
        cell = self.__ACTIVE_COLUMNS[1] + str(presentation_row)
        border = Border(bottom=Side(border_style='thin'))
        sheet[cell].font = Font(size=self.LARGE_FONT_SIZE)
        sheet[cell].alignment = Alignment(horizontal='left', vertical='center')
        sheet[cell].border = border
        # The last two cells of the row must also be underlined (where presenter names go)
        # Usually, its only the last cell that is not underlined. Only the presentations
        # under the IMPROVE_IN_MINISTRY section leave the last two cells not underlined.
        # But, a general solution is to underline the last two cells of all presentation rows
        cell = self.__ACTIVE_COLUMNS[2] + str(presentation_row)
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(presentation_row)
        sheet[cell].border = border

    def _insert_footer_content(self, sheet: Worksheet):  # todo: get values from config file
        footer_row = self.__CURRENT_ROW
        current_row = str(self.__CURRENT_ROW)
        cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[cell] = 'Reader'
        self.__CURRENT_ROW += 1

        current_row = str(self.__CURRENT_ROW)
        cell = self.__ACTIVE_COLUMNS[2] + current_row
        sheet[cell] = 'Concluding Prayer'
        self.__CURRENT_ROW += 3
        self._style_footer_content(footer_row, sheet)

    def _style_footer_content(self, footer_row: int, sheet: Worksheet):
        font = Font(size=self.SMALL_FONT_SIZE)
        alignment = Alignment(horizontal='left', vertical='center')
        border = Border(bottom=Side(border_style='thin'))

        cell = self.__ACTIVE_COLUMNS[2] + str(footer_row)
        sheet[cell].font = font
        sheet[cell].alignment = alignment
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(footer_row)
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[2] + str(footer_row + 1)
        sheet[cell].font = font
        sheet[cell].alignment = alignment
        sheet[cell].border = border
        cell = self.__ACTIVE_COLUMNS[3] + str(footer_row + 1)
        sheet[cell].border = border

    def _final_styling_touches(self, sheet: Worksheet):
        sheet.page_setup.paperSize = Worksheet.PAPERSIZE_A4
        sheet.column_dimensions[self.__LEFT_COLUMNS[0]].width = 4
        sheet.column_dimensions[self.__RIGHT_COLUMNS[0]].width = 4
